#!/usr/bin/env python3
# -*- coding:utf-8 -*-
'''
Copyright (c) 2024, Yehua Shang

  @Author: Yehua Shang
  @Email:  niceshang@outlook.com
  @Time:   2024-11-29 17:11:33

    将数据库差异信息导出到excel文件

'''
import json

from openpyxl import load_workbook

# 保存数据库信息
def write_table_info(sheet,db_info):
    sheet.cell(row=1,column=1,value=f'数据库信息文件：{db_info["file_path"]}')
    tables = db_info['tables']
    for i,table in enumerate(tables):
        sheet.cell(row=3+i,column=1,value=table['v_name'])
        sheet.cell(row=3+i,column=2,value=table['count'])
        sheet.cell(row=3+i,column=3,value=table['ddl'])

# 保存表名和表数量
def write_table_count(sheet,db_info,table_list):
    if not table_list:
        return
    tables = db_info['tables']
    name_set = set(table_list)
    row_cursor = 1
    for table in tables:
        if table['v_name'] in name_set:
            row_cursor += 1
            sheet.cell(row=row_cursor,column=1,value=table['v_name'])
            sheet.cell(row=row_cursor,column=2,value=table['count'])

# 保存表差异
def write_table_diff(sheet,diff_tables):
    if not diff_tables:
        return

    tbl_cursor = 3
    col_cursor = 3
    idx_cursor = 3
    fk_cursor = 3

    for table in diff_tables:
        sheet.cell(row=tbl_cursor,column=1,value=table['table_name'])
        tbl_cursor += 1

        # 字段差异
        diff_columns = table['diff_columns']
        if diff_columns:
            sheet.cell(row=col_cursor,column=2,value='字段差异')
            for column in diff_columns:
                sheet.cell(row=col_cursor,column=2,value=column['col_name'])
                sheet.cell(row=col_cursor,column=3,value=column['col1_def'])
                sheet.cell(row=col_cursor,column=4,value=column['col2_def'])
                col_cursor += 1
        
        # 索引差异
        diff_indices = table['diff_indices']
        if diff_indices:
            for index in diff_indices:
                sheet.cell(row=idx_cursor,column=5,value=index['index_name'])
                sheet.cell(row=idx_cursor,column=6,value=index['index1_def'])
                sheet.cell(row=idx_cursor,column=7,value=index['index2_def'])
                idx_cursor += 1

        # 外键差异
        diff_fks = table['diff_foreign_keys']
        if diff_fks:
            for fk in diff_fks:
                sheet.cell(row=fk_cursor,column=8,value=fk['fk_name'])
                sheet.cell(row=fk_cursor,column=9,value=fk['fk1_def'])
                sheet.cell(row=fk_cursor,column=10,value=fk['fk2_def'])
                fk_cursor += 1
        
        # 对齐各个cursor
        tbl_cursor = max(tbl_cursor,col_cursor,idx_cursor,fk_cursor)
        col_cursor = tbl_cursor
        idx_cursor = tbl_cursor
        fk_cursor = tbl_cursor

# 写出视图信息
def write_view_info(sheet,db_info,view_list):
    if not view_list:
        return
    views = db_info['views']
    if not views:
        return
    row_cursor = 1
    for view in views:
        row_cursor += 1
        sheet.cell(row=row_cursor,column=1,value=view['v_name'])
        sheet.cell(row=row_cursor,column=2,value=view['ddl'])

# 写出视图差异
def write_view_diff(sheet,diff_views):
    if not diff_views:
        return
    for i,view in enumerate(diff_views):
        sheet.cell(row=i+3,column=1,value=view['view_name'])
        sheet.cell(row=i+3,column=2,value=view['view1_ddl'])
        sheet.cell(row=i+3,column=3,value=view['view2_ddl'])
        
    
# 写出存储过程信息
def write_proc_info(sheet,db_info,proc_list):
    if not proc_list:
        return
    procs = db_info['procedures']
    if not procs:
        return
    row_cursor = 1
    for proc in procs:
        row_cursor += 1
        sheet.cell(row=row_cursor,column=1,value=proc['name'])
        sheet.cell(row=row_cursor,column=2,value=proc['definition'])

# 写出存储过程差异
def write_proc_diff(sheet,diff_procs):
    if not diff_procs:
        return
    for i,proc in enumerate(diff_procs):
        sheet.cell(row=i+3,column=1,value=proc['procedure_name'])
        sheet.cell(row=i+3,column=2,value=proc['procedure1_def'])
        sheet.cell(row=i+3,column=3,value=proc['procedure2_def'])

# 写出文件差异信息
def write_diff(template_file,diff_info,db1,db2,diff_file):
    template_wb = load_workbook(template_file)
    # 写出数据库信息
    sheet = template_wb['t_in_db1']
    write_table_info(sheet,db1)
    sheet = template_wb['t_in_db2']
    write_table_info(sheet,db2)

    # 写出仅在db1中的表
    sheet = template_wb['t_o_in_db1']
    write_table_count(sheet,db1,diff_info['tables_only_in_db1'])
    # 写出仅在db2中的表
    sheet = template_wb['t_o_in_db2']
    write_table_count(sheet,db2,diff_info['tables_only_in_db2'])

    # 写出相同的表
    sheet = template_wb['cmn_t']
    write_table_count(sheet,db1,diff_info['common_tables'])

    # 写出表差异
    sheet = template_wb['dif_t']
    write_table_diff(sheet,diff_info['diff_tables'])

    # 写出视图信息
    sheet = template_wb['v_o_in_db1']
    write_view_info(sheet,db1,diff_info['views_only_in_db1'])
    sheet = template_wb['v_o_in_db2']
    write_view_info(sheet,db2,diff_info['views_only_in_db2'])
    sheet = template_wb['cmn_v']
    write_view_info(sheet,db1,diff_info['common_views'])
    sheet = template_wb['dif_v']
    write_view_diff(sheet,diff_info['diff_views'])

    # 写出存储过程信息
    sheet = template_wb['p_o_in_db1']
    write_proc_info(sheet,db1,diff_info['procedures_only_in_db1'])
    sheet = template_wb['p_o_in_db2']
    write_proc_info(sheet,db2,diff_info['procedures_only_in_db2'])
    sheet = template_wb['cmn_p']
    write_proc_info(sheet,db1,diff_info['common_procedures'])
    sheet = template_wb['dif_p']
    write_proc_diff(sheet,diff_info['diff_procedures'])

    template_wb.save(diff_file)

def read_db_file(file_path):
    # 读取 JSON 文件
    with open(file_path, "r", encoding="utf-8") as file:
        data = json.load(file)
        data["file_path"] = file_path

    return data

if __name__ == '__main__':
    import os
    file_dir = '/home/sany/tmp/omnidb-files/'
    db1 = read_db_file(file_dir + "my_test_24-11-28-110158_00.json")
    db2 = read_db_file(file_dir + "my_test_24-11-28-112640_01.json")
    diff_info = read_db_file(file_dir + "my_test_1129_diff.json")

    base_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_template_file = base_dir + '/templates/db_info_diff.xlsx'
    diff_file = base_dir + '/omnidb-files/diff_result.xlsx'

    write_diff(xlsx_template_file,diff_info,db1,db2,diff_file)
